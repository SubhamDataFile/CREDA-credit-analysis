import pdfplumber
import re
import math
from openpyxl import load_workbook



RISK_THRESHOLDS = {
    "Current Ratio": {"green": 1.5, "amber": 1.0},
    "Debt-Equity Ratio": {"green": 1.0, "amber": 2.0},
    "Interest Coverage Ratio": {"green": 3.0, "amber": 1.5},
    "DSCR": {"green": 1.25, "amber": 1.0},
    "ROCE": {"green": 0.15, "amber": 0.10},
    "ROA": {"green": 0.08, "amber": 0.04},
    "EBITDA Margin": {"green": 0.20, "amber": 0.12},
    "Net Profit Margin": {"green": 0.10, "amber": 0.05}
}



PL_LABELS = {
    "Revenue": ["revenue"],
    "Net Profit": ["profit for the year"],
    "PBT": ["profit before tax"],
    "Interest Expense": ["finance cost"],
    "Depreciation": ["depreciation"]
}

BS_LABELS = {
    "Current Assets": ["total current assets"],
    "Current Liabilities": ["total current liabilities"],
    "Total Assets": ["total assets"],
    "Net Worth": ["total equity"]
}



def parse_number(x):
    x = x.replace(",", "")
    if x.startswith("(") and x.endswith(")"):
        return -float(x[1:-1])
    return float(x)


def assign_risk_flag(value, metric):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "NA"
    t = RISK_THRESHOLDS.get(metric)
    if not t:
        return "NA"
    if value >= t["green"]:
        return "GREEN"
    elif value >= t["amber"]:
        return "AMBER"
    return "RED"



def find_consolidated_section(pdf):
    for i, page in enumerate(pdf.pages):
        txt = page.extract_text()
        if txt and "consolidated financial statements" in txt.lower():
            return i
    return None


def extract_rupee_crore_tables(pdf, start_page, max_pages=40):
    """
    Extract ONLY tables whose header contains '₹ crore' or 'rs. crore'
    """
    valid_tables = []

    for p in range(start_page, min(start_page + max_pages, len(pdf.pages))):
        page = pdf.pages[p]
        text = page.extract_text()
        if not text:
            continue

        header = text.lower()
        if "₹ crore" not in header and "rs. crore" not in header:
            continue

        tables = page.extract_tables()
        if tables:
            valid_tables.extend(tables)

    return valid_tables


def extract_from_tables(tables, label_map):
    results = {}

    for table in tables:
        for row in table:
            if not row or len(row) < 2:
                continue

            label = str(row[0]).lower().strip()

            for metric, keywords in label_map.items():
                if any(k in label for k in keywords):
                    val = row[-1]
                    if isinstance(val, str) and re.search(r"\d", val):
                        results[metric] = parse_number(val)
                        break

    return results



def run_financial_analysis(pdf_path):

    financial_data = {}

    with pdfplumber.open(pdf_path) as pdf:

        
        start_page = find_consolidated_section(pdf)
        if start_page is None:
            raise ValueError("Consolidated Financial Statements section not found")

        financial_data["Statement Type"] = "Consolidated (₹ crore)"

        
        tables = extract_rupee_crore_tables(pdf, start_page)

        if not tables:
            raise ValueError("No ₹ crore consolidated tables found")

        
        pl_data = extract_from_tables(tables, PL_LABELS)
        bs_data = extract_from_tables(tables, BS_LABELS)

        financial_data.update(pl_data)
        financial_data.update(bs_data)

        
        financial_data["Total Debt"] = 0.0

    
    financial_data["EBIT"] = financial_data.get("PBT", 0) + financial_data.get("Interest Expense", 0)
    financial_data["EBITDA"] = financial_data["EBIT"] + financial_data.get("Depreciation", 0)
    financial_data["Principal Repayment"] = 0.0

    
    ca = financial_data.get("Current Assets", 0)
    cl = financial_data.get("Current Liabilities", 0)
    nw = financial_data.get("Net Worth", 0)
    td = financial_data.get("Total Debt", 0)
    ta = financial_data.get("Total Assets", nw + td)
    rev = financial_data.get("Revenue", 0)
    np = financial_data.get("Net Profit", 0)
    ebit = financial_data.get("EBIT", 0)
    ebitda = financial_data.get("EBITDA", 0)
    interest = financial_data.get("Interest Expense", 0)

    capital_employed = nw + td if nw > 0 else 0

    financial_data["Current Ratio"] = ca / cl if cl > 0 else 0
    financial_data["Quick Ratio"] = financial_data["Current Ratio"]
    financial_data["Debt-Equity Ratio"] = 0
    financial_data["Equity Ratio"] = nw / ta if ta > 0 else 0
    financial_data["Interest Coverage Ratio"] = ebit / interest if interest > 0 else 0
    financial_data["DSCR"] = ebitda / interest if interest > 0 else 0
    financial_data["ROCE"] = ebit / capital_employed if capital_employed > 0 else 0
    financial_data["ROA"] = np / ta if ta > 0 else 0
    financial_data["EBITDA Margin"] = ebitda / rev if rev > 0 else 0
    financial_data["Net Profit Margin"] = np / rev if rev > 0 else 0

    
    ratio_results = {k: financial_data.get(k) for k in RISK_THRESHOLDS}
    risk_flags = {k: assign_risk_flag(v, k) for k, v in ratio_results.items()}

    
    wb = load_workbook("AI output.xlsx")

    ws_fin = wb["financial_data"]
    for row in ws_fin.iter_rows(min_row=2, max_col=2):
        if row[0].value in financial_data:
            row[1].value = financial_data[row[0].value]

    ws_rat = wb["Ratio_Calculations"]
    for row in ws_rat.iter_rows(min_row=2):
        if row[0].value in ratio_results:
            row[1].value = ratio_results[row[0].value]

    if "Credit_Risk_Flags" not in wb.sheetnames:
        ws_flag = wb.create_sheet("Credit_Risk_Flags")
        ws_flag.append(["Metric", "Value", "Risk Flag"])
    else:
        ws_flag = wb["Credit_Risk_Flags"]
        ws_flag.delete_rows(2, ws_flag.max_row)

    for k, v in ratio_results.items():
        ws_flag.append([k, v, risk_flags[k]])

    wb.save("AI output.xlsx")
